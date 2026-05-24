#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Watermark Removal Demo
======================
Reads syria_cars_with_images_sample.xlsx, downloads real car images,
removes the syriacars semi-transparent text watermark using LaMa inpainting,
and saves side-by-side before/after comparisons.

USAGE
-----
  python watermark_demo.py
  python watermark_demo.py --max 5          # process first 5 cars only
  python watermark_demo.py --source damazzle
"""

import argparse
import io
import os
import random
import sys
import time
from pathlib import Path
from typing import Optional

import cv2
import numpy as np
import openpyxl
import requests
from PIL import Image, ImageDraw, ImageFont

sys.stdout.reconfigure(encoding='utf-8')

# ── Config ────────────────────────────────────────────────────────────────────

EXCEL_PATH   = r'C:\Users\Tasnaim\Downloads\syria_cars_with_images_sample.xlsx'
OUTPUT_DIR   = Path(r'D:\sky_scraper\skyscraper\watermark_demo_output_v2')

HEADERS = {
    'User-Agent': random.choice([
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/119.0.0.0',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/120.0.0.0',
    ]),
    'Accept': 'image/webp,image/apng,image/*,*/*;q=0.8',
    'Referer': 'https://syriacars.net/',
}


# ── Watermark detection ───────────────────────────────────────────────────────

def detect_syriacars_mask(img_bgr: np.ndarray) -> np.ndarray:
    """
    Syriacars watermark mask — tight fixed rectangle applied to ALL cars.

    The syriacars watermark is centred at ~50 % height regardless of car colour.
    We use a narrow 24 % wide × 7 % tall rectangle (1.7 % of image area).
    LaMa handles this small region cleanly on both white and coloured cars.

    Coverage comparison:
      Old mask  → 28 % × 20 % = 5.6 % of image (too destructive)
      New mask  → 24 % × 7 %  = 1.7 % of image (tight, LaMa handles well)
    """
    h, w = img_bgr.shape[:2]

    mw = int(w * 0.24)
    mh = int(h * 0.07)
    cx = w // 2
    cy = h // 2
    x0 = max(0, cx - mw // 2);  x1 = min(w, x0 + mw)
    y0 = max(0, cy - mh // 2);  y1 = min(h, y0 + mh)
    mask = np.zeros((h, w), dtype=np.uint8)
    mask[y0:y1, x0:x1] = 255
    coverage = mask.mean() / 255 * 100
    print(f' [mask {x1-x0}×{y1-y0}px = {coverage:.1f}%]', end='')
    return mask


def detect_damazzle_mask(img_bgr: np.ndarray) -> np.ndarray:
    """
    Detect the damazzle orange/red semi-transparent 'Z' logo.
    Uses HSV colour thresholding for the orange range.
    Returns a uint8 mask (255 = inpaint this region, 0 = keep).
    """
    hsv = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2HSV)

    # Orange hue range (OpenCV H is 0-179)
    lower1 = np.array([0,  80,  80])
    upper1 = np.array([20, 255, 255])
    lower2 = np.array([160, 80, 80])
    upper2 = np.array([179, 255, 255])

    mask1 = cv2.inRange(hsv, lower1, upper1)
    mask2 = cv2.inRange(hsv, lower2, upper2)
    mask  = cv2.bitwise_or(mask1, mask2)

    # Clean up
    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (7, 7))
    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel)
    mask = cv2.dilate(mask, kernel, iterations=2)

    return mask


def get_mask(img_bgr: np.ndarray, source: str = 'syriacars') -> np.ndarray:
    if source == 'damazzle':
        return detect_damazzle_mask(img_bgr)
    return detect_syriacars_mask(img_bgr)


# ── Inpainting ────────────────────────────────────────────────────────────────

_lama_model = None

def load_lama():
    """Load LaMa model once and cache it (uses simple-lama-inpainting)."""
    global _lama_model
    if _lama_model is None:
        print('  Loading LaMa model (first run downloads ~500 MB) …')
        from simple_lama_inpainting import SimpleLama
        _lama_model = SimpleLama()
        print('  LaMa model ready.')
    return _lama_model


def inpaint_lama(img_bgr: np.ndarray, mask: np.ndarray) -> np.ndarray:
    """
    Run LaMa inpainting via simple-lama-inpainting.
    img_bgr : H×W×3 uint8 BGR
    mask    : H×W uint8  (255 = region to fill)
    Returns inpainted H×W×3 uint8 BGR.
    """
    model = load_lama()
    # SimpleLama expects PIL RGB image + PIL grayscale mask
    img_rgb  = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB)
    pil_img  = Image.fromarray(img_rgb)
    pil_mask = Image.fromarray(mask)
    result_pil = model(pil_img, pil_mask)
    return cv2.cvtColor(np.array(result_pil), cv2.COLOR_RGB2BGR)


def inpaint_opencv_fallback(img_bgr: np.ndarray, mask: np.ndarray) -> np.ndarray:
    """Fast OpenCV TELEA inpaint — used if LaMa is unavailable."""
    return cv2.inpaint(img_bgr, mask, inpaintRadius=7, flags=cv2.INPAINT_TELEA)


def remove_watermark(img_bgr: np.ndarray, mask: np.ndarray,
                     use_lama: bool = True) -> np.ndarray:
    if mask.max() == 0:
        return img_bgr   # nothing detected
    if use_lama:
        try:
            return inpaint_lama(img_bgr, mask)
        except Exception as exc:
            print(f'  LaMa failed ({exc}), falling back to OpenCV inpaint.')
    return inpaint_opencv_fallback(img_bgr, mask)


# ── Comparison image ──────────────────────────────────────────────────────────

def make_comparison(before: np.ndarray, after: np.ndarray,
                    mask: np.ndarray, label: str) -> np.ndarray:
    """Create a side-by-side: before | mask (red overlay) | after."""
    h, w = before.shape[:2]
    target_w = 600
    scale = target_w / w
    nh, nw = int(h * scale), target_w

    b = cv2.resize(before, (nw, nh))
    a = cv2.resize(after,  (nw, nh))
    m = cv2.resize(mask,   (nw, nh))

    # Red overlay on before to show the mask
    overlay = b.copy()
    overlay[m > 127] = [0, 0, 200]
    masked_vis = cv2.addWeighted(b, 0.6, overlay, 0.4, 0)

    # Labels
    def put_label(img, text):
        return cv2.putText(img.copy(), text, (8, 24),
                           cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)

    b_l = put_label(b,           'BEFORE')
    m_l = put_label(masked_vis,  'MASK (red)')
    a_l = put_label(a,           'AFTER')

    combined = np.hstack([b_l, m_l, a_l])

    # Title bar
    title_bar = np.zeros((36, combined.shape[1], 3), dtype=np.uint8)
    cv2.putText(title_bar, label, (8, 26),
                cv2.FONT_HERSHEY_SIMPLEX, 0.65, (220, 220, 220), 1)

    return np.vstack([title_bar, combined])


# ── Image download ────────────────────────────────────────────────────────────

def download_image(url: str) -> Optional[np.ndarray]:
    try:
        r = requests.get(url.strip(), headers=HEADERS, timeout=20)
        r.raise_for_status()
        arr = np.frombuffer(r.content, np.uint8)
        img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
        return img
    except Exception as exc:
        print(f'    Download failed: {exc}')
        return None


# ── Main demo ─────────────────────────────────────────────────────────────────

def run_demo(source: str = 'syriacars', max_cars: int = 5, use_lama: bool = True):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    before_dir = OUTPUT_DIR / 'before'
    after_dir  = OUTPUT_DIR / 'after'
    cmp_dir    = OUTPUT_DIR / 'comparison'
    for d in [before_dir, after_dir, cmp_dir]:
        d.mkdir(exist_ok=True)

    # ── Read Excel ─────────────────────────────────────────────────────────
    print(f'Reading {EXCEL_PATH} …')
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    orig_idx = headers.index('images_original_links')
    id_idx   = headers.index('id')
    src_idx  = headers.index('source')

    cars = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        s = (row[src_idx] or '').strip()
        links = (row[orig_idx] or '').strip()
        if s == source and links:
            cars.append({'id': row[id_idx], 'links': links})
        if len(cars) >= max_cars:
            break
    wb.close()
    print(f'Found {len(cars)} {source} cars with images. Processing …\n')

    # ── Process each car ───────────────────────────────────────────────────
    stats = {'total_imgs': 0, 'detected': 0, 'cleaned': 0, 'failed': 0}

    for car in cars:
        car_id   = str(car['id']).replace('/', '_')
        img_urls = [u.strip() for u in car['links'].split(',') if u.strip()]
        print(f'[{car_id}] — {len(img_urls)} image(s)')

        for i, url in enumerate(img_urls, 1):
            stats['total_imgs'] += 1
            fname = f'{car_id}_{i:02d}'
            print(f'  Img {i}: {url[-60:]}')

            # Download
            img = download_image(url)
            if img is None:
                stats['failed'] += 1
                time.sleep(1)
                continue

            # Save before
            cv2.imwrite(str(before_dir / f'{fname}_before.jpg'), img)

            # Detect watermark
            mask = get_mask(img, source=source)
            coverage = (mask > 0).sum() / mask.size * 100
            print(f'  Mask coverage: {coverage:.1f}%', end='')

            if coverage < 0.1:
                print(' — no watermark detected, skipping')
                stats['failed'] += 1
                time.sleep(0.8)
                continue

            stats['detected'] += 1
            print(f' — running {"LaMa" if use_lama else "OpenCV"} inpaint …', end='', flush=True)

            # Remove watermark
            cleaned = remove_watermark(img, mask, use_lama=use_lama)
            print(' done')
            stats['cleaned'] += 1

            # Save after + comparison
            cv2.imwrite(str(after_dir  / f'{fname}_after.jpg'), cleaned)
            cmp = make_comparison(img, cleaned, mask, f'{car_id} | img {i}')
            cv2.imwrite(str(cmp_dir    / f'{fname}_cmp.jpg'), cmp)

            time.sleep(random.uniform(1.0, 2.5))

        print()

    # ── Summary ────────────────────────────────────────────────────────────
    print('=' * 56)
    print(f'Done.')
    print(f'  Images downloaded : {stats["total_imgs"]}')
    print(f'  Watermark detected: {stats["detected"]}')
    print(f'  Successfully cleaned: {stats["cleaned"]}')
    print(f'  Failed/skipped    : {stats["failed"]}')
    print(f'\nOutput saved to:')
    print(f'  Before      : {before_dir}')
    print(f'  After       : {after_dir}')
    print(f'  Comparisons : {cmp_dir}')
    print('=' * 56)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Watermark removal demo')
    parser.add_argument('--max',    type=int, default=5, metavar='N',
                        help='Max number of cars to process (default 5)')
    parser.add_argument('--source', default='syriacars',
                        choices=['syriacars', 'damazzle'],
                        help='Which site watermark to target')
    parser.add_argument('--opencv', action='store_true',
                        help='Use OpenCV inpaint instead of LaMa (faster but lower quality)')
    args = parser.parse_args()

    run_demo(source=args.source, max_cars=args.max, use_lama=not args.opencv)
