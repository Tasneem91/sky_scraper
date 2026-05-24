#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Damazzle Watermark Removal Demo
=================================
Reads Damazzle (1).xlsx, downloads original car images,
detects the golden/amber "ZS" logo watermark by color (HSV),
removes it with LaMa inpainting, and saves side-by-side comparisons.

USAGE
-----
  python watermark_demo_damazzle.py              # 5 cars
  python watermark_demo_damazzle.py --max 10     # 10 cars
  python watermark_demo_damazzle.py --opencv     # faster, lower quality
"""

import argparse
import os
import random
import sys
import time
from pathlib import Path
from typing import Optional

import cv2
import numpy as np
import openpyxl
import requests
from PIL import Image

sys.stdout.reconfigure(encoding='utf-8')

# ── Config ────────────────────────────────────────────────────────────────────

EXCEL_PATH = r'C:\Users\Tasnaim\Downloads\Damazzle (1).xlsx'
OUTPUT_DIR = Path(r'D:\sky_scraper\skyscraper\watermark_demo_damazzle')

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0',
    'Accept': 'image/webp,image/apng,image/*,*/*;q=0.8',
    'Referer': 'https://damazzle.com/',
}


# ── Watermark detection ───────────────────────────────────────────────────────

def detect_damazzle_mask(img_bgr: np.ndarray) -> np.ndarray:
    """
    Detect the Damazzle golden/amber watermark logo by HSV color + position filtering.

    Strategy:
      1. Color mask   — gold/amber hue (H 8-38 in OpenCV 0-179 scale)
      2. Blob filter  — keep only blobs that are:
                          • small enough  (< 8% of image — rejects sky/large backgrounds)
                          • large enough  (> 0.05% of image — rejects noise)
                          • centroid inside center 60% of image (rejects edge headlights,
                            taillights, gear knobs, sky at top)
      3. Best blob    — of remaining candidates, keep the one closest to image center
                        plus any adjacent blob within 15% of image size (logo = 2 parts)
      4. Dilation     — expand to cover semi-transparent watermark edges
    """
    h, w = img_bgr.shape[:2]
    cx, cy = w // 2, h // 2
    hsv = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2HSV)

    # Gold/amber hue range
    lower = np.array([8,  100, 100], dtype=np.uint8)
    upper = np.array([38, 255, 255], dtype=np.uint8)
    color_mask = cv2.inRange(hsv, lower, upper)

    # Close small gaps before labelling
    kernel_close = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (7, 7))
    color_mask = cv2.morphologyEx(color_mask, cv2.MORPH_CLOSE, kernel_close)

    # Connected components
    nb, labels, stats, centroids = cv2.connectedComponentsWithStats(
        color_mask, connectivity=8
    )

    min_area   = int(h * w * 0.0005)   # 0.05% — reject noise
    max_area   = int(h * w * 0.08)     # 8%    — reject sky / huge backgrounds
    x_margin   = int(w * 0.20)         # centre 60% horizontally
    y_margin   = int(h * 0.20)         # centre 60% vertically

    # Find best blob (closest to centre, within constraints)
    best_lbl  = None
    best_dist = float('inf')
    for lbl in range(1, nb):
        area = stats[lbl, cv2.CC_STAT_AREA]
        bx, by = centroids[lbl]
        if area < min_area or area > max_area:
            continue
        if not (x_margin < bx < w - x_margin and y_margin < by < h - y_margin):
            continue
        dist = ((bx - cx) ** 2 + (by - cy) ** 2) ** 0.5
        if dist < best_dist:
            best_dist = dist
            best_lbl  = lbl

    if best_lbl is None:
        print(' [no watermark detected]', end='')
        return np.zeros((h, w), dtype=np.uint8)

    best_cx, best_cy = centroids[best_lbl]
    proximity = max(w, h) * 0.15   # include nearby blobs (logo has two parts)

    final_mask = np.zeros((h, w), dtype=np.uint8)
    for lbl in range(1, nb):
        area = stats[lbl, cv2.CC_STAT_AREA]
        if area < min_area or area > max_area:
            continue
        bx, by = centroids[lbl]
        dist_to_best = ((bx - best_cx) ** 2 + (by - best_cy) ** 2) ** 0.5
        if dist_to_best <= proximity:
            final_mask[labels == lbl] = 255

    # Dilate generously to cover semi-transparent edges (especially on dark surfaces)
    kernel_dilate = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (17, 17))
    final_mask = cv2.dilate(final_mask, kernel_dilate, iterations=3)

    coverage = final_mask.mean() / 255 * 100
    print(f' [mask {coverage:.1f}% of image]', end='')
    return final_mask


# ── Inpainting ────────────────────────────────────────────────────────────────

_lama_model = None

def load_lama():
    global _lama_model
    if _lama_model is None:
        print('  Loading LaMa model …')
        from simple_lama_inpainting import SimpleLama
        _lama_model = SimpleLama()
        print('  LaMa ready.')
    return _lama_model


def inpaint_lama(img_bgr: np.ndarray, mask: np.ndarray) -> np.ndarray:
    model = load_lama()
    img_rgb  = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB)
    pil_img  = Image.fromarray(img_rgb)
    pil_mask = Image.fromarray(mask)
    result   = model(pil_img, pil_mask)
    return cv2.cvtColor(np.array(result), cv2.COLOR_RGB2BGR)


def inpaint_opencv(img_bgr: np.ndarray, mask: np.ndarray) -> np.ndarray:
    return cv2.inpaint(img_bgr, mask, inpaintRadius=7, flags=cv2.INPAINT_TELEA)


def remove_watermark(img_bgr: np.ndarray, mask: np.ndarray,
                     use_lama: bool = True) -> np.ndarray:
    if mask.max() == 0:
        return img_bgr
    if use_lama:
        try:
            return inpaint_lama(img_bgr, mask)
        except Exception as exc:
            print(f'  LaMa failed ({exc}), falling back to OpenCV.')
    return inpaint_opencv(img_bgr, mask)


# ── Comparison image ──────────────────────────────────────────────────────────

def make_comparison(before: np.ndarray, after: np.ndarray,
                    mask: np.ndarray, label: str) -> np.ndarray:
    target_w = 640
    scale = target_w / before.shape[1]
    nh = int(before.shape[0] * scale)

    b = cv2.resize(before, (target_w, nh))
    a = cv2.resize(after,  (target_w, nh))
    m = cv2.resize(mask,   (target_w, nh))

    overlay = b.copy()
    overlay[m > 127] = [0, 0, 200]
    masked_vis = cv2.addWeighted(b, 0.55, overlay, 0.45, 0)

    def label_img(img, text):
        out = img.copy()
        cv2.putText(out, text, (8, 28), cv2.FONT_HERSHEY_SIMPLEX,
                    0.8, (0, 0, 0),   3, cv2.LINE_AA)
        cv2.putText(out, text, (8, 28), cv2.FONT_HERSHEY_SIMPLEX,
                    0.8, (255, 255, 255), 1, cv2.LINE_AA)
        return out

    combined = np.hstack([
        label_img(b,          'BEFORE'),
        label_img(masked_vis, 'MASK'),
        label_img(a,          'AFTER'),
    ])

    title = np.full((38, combined.shape[1], 3), 30, dtype=np.uint8)
    cv2.putText(title, label, (8, 26), cv2.FONT_HERSHEY_SIMPLEX,
                0.65, (200, 200, 200), 1, cv2.LINE_AA)
    return np.vstack([title, combined])


# ── Image download ────────────────────────────────────────────────────────────

def download_image(url: str) -> Optional[np.ndarray]:
    try:
        r = requests.get(url.strip(), headers=HEADERS, timeout=25)
        r.raise_for_status()
        arr = np.frombuffer(r.content, np.uint8)
        img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
        return img
    except Exception as exc:
        print(f'    Download failed: {exc}')
        return None


# ── Main demo ─────────────────────────────────────────────────────────────────

def run_demo(max_cars: int = 5, use_lama: bool = True):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    before_dir = OUTPUT_DIR / 'before'
    after_dir  = OUTPUT_DIR / 'after'
    cmp_dir    = OUTPUT_DIR / 'comparison'
    for d in [before_dir, after_dir, cmp_dir]:
        d.mkdir(exist_ok=True)

    # ── Read Excel ─────────────────────────────────────────────────────────
    print(f'Reading {EXCEL_PATH} …')
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    orig_col = headers.index('images_original_links')
    id_col   = headers.index('id')

    cars = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        links = (row[orig_col] or '').strip()
        if links:
            cars.append({'id': row[id_col], 'links': links})
        if len(cars) >= max_cars:
            break
    wb.close()
    print(f'Found {len(cars)} cars. Processing …\n')

    stats = {'total': 0, 'detected': 0, 'cleaned': 0, 'failed': 0}

    for car in cars:
        car_id   = str(car['id']).replace('/', '_')
        img_urls = [u.strip() for u in car['links'].split(',') if u.strip()]
        print(f'[{car_id}] — {len(img_urls)} image(s)')

        for i, url in enumerate(img_urls, 1):
            stats['total'] += 1
            fname = f'{car_id}_{i:02d}'
            print(f'  Img {i}: …{url[-55:]}', end='')

            img = download_image(url)
            if img is None:
                stats['failed'] += 1
                print()
                time.sleep(1)
                continue

            cv2.imwrite(str(before_dir / f'{fname}_before.jpg'), img)

            mask = detect_damazzle_mask(img)
            coverage = mask.mean() / 255 * 100

            if coverage < 0.1:
                print(' — no watermark detected, skipping')
                stats['failed'] += 1
                time.sleep(0.8)
                continue

            stats['detected'] += 1
            method = 'LaMa' if use_lama else 'OpenCV'
            print(f' — {method} inpaint …', end='', flush=True)

            cleaned = remove_watermark(img, mask, use_lama=use_lama)
            print(' done')
            stats['cleaned'] += 1

            cv2.imwrite(str(after_dir / f'{fname}_after.jpg'), cleaned)
            cmp = make_comparison(img, cleaned, mask, f'{car_id} | img {i}')
            cv2.imwrite(str(cmp_dir   / f'{fname}_cmp.jpg'), cmp)

            time.sleep(random.uniform(1.0, 2.0))

        print()

    print('=' * 56)
    print(f'Done.')
    print(f'  Downloaded  : {stats["total"]}')
    print(f'  Detected    : {stats["detected"]}')
    print(f'  Cleaned     : {stats["cleaned"]}')
    print(f'  Failed/skip : {stats["failed"]}')
    print(f'\nResults in: {OUTPUT_DIR}')
    print('=' * 56)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Damazzle watermark removal demo')
    parser.add_argument('--max',    type=int, default=5, metavar='N',
                        help='Max cars to process (default 5)')
    parser.add_argument('--opencv', action='store_true',
                        help='Use OpenCV inpaint instead of LaMa (faster, lower quality)')
    args = parser.parse_args()

    run_demo(max_cars=args.max, use_lama=not args.opencv)
